import openpyxl as xl
from openpyxl.chart import BarChart, Reference


def process_sheet(filename):
    wb = xl.load_workbook(filename)
    sheet = wb.get_sheet_by_name('Sheet1')

    sheet.cell(1, 4).value = 'discount_price'
    for row in range(2, sheet.max_row + 1):
        price = sheet.cell(row, 3).value
        discount_price = price * 0.9
        sheet.cell(row, 4).value = discount_price

    values = Reference(sheet, min_row=2, max_row=sheet.max_row, min_col=4, max_col=4)
    chart = BarChart()
    chart.add_data(values, titles_from_data=False)
    chart.title = "Discount price"
    chart.x_axis.title = "Item"
    chart.y_axis.title ="Price (after discount)"
    categories = Reference(sheet, min_row=2, max_row=sheet.max_row, min_col=2, max_col=2)
    chart.set_categories(categories)
    sheet.add_chart(chart, 'e2')
    wb.save(filename)